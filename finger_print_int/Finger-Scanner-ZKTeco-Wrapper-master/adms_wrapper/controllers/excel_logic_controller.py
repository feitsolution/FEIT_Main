import io
from datetime import datetime, time, timedelta
from typing import Any

import pandas as pd
from openpyxl.styles import PatternFill

from adms_wrapper.controllers.data_processing_controller import process_attendance_summary
from adms_wrapper.models.db_queries_models import (
    get_default_shift,
    get_device_branch_mappings,
    get_employee_branch_mappings,
    get_employee_designation_mappings,
    get_employee_name_mappings,
    get_setting,
    get_shift_templates,
)

# ==============================================================================
# Constants
# ==============================================================================

NOON_HOUR = 12

# ==============================================================================
# Mapping Functions
# ==============================================================================


def map_branch(sn: Any, mappings_df: pd.DataFrame) -> str:
    if pd.isna(sn) or sn is None:
        return ""

    if mappings_df is None or mappings_df.empty:
        return ""

    df = mappings_df.copy()

    if "serial_number" not in df.columns or "branch_name" not in df.columns:
        rename_map = {}
        for alt in ("serial", "sn", "device_sn", "serial_no"):
            if alt in df.columns and "serial_number" not in df.columns:
                rename_map[alt] = "serial_number"
                break
        for alt in ("branch", "branchname", "branch_name_str"):
            if alt in df.columns and "branch_name" not in df.columns:
                rename_map[alt] = "branch_name"
                break
        if rename_map:
            df = df.rename(columns=rename_map)

    if "serial_number" not in df.columns or "branch_name" not in df.columns:
        return ""

    try:
        sn_str = str(sn).strip()
        match = df[df["serial_number"].astype(str).str.strip() == sn_str]
    except Exception:
        try:
            match = df[df["serial_number"].astype(str).str.contains(str(sn), na=False)]
        except Exception:
            return ""

    if not match.empty:
        return match.iloc[0]["branch_name"]
    return ""


def map_designation(emp_id: Any, designation_df: pd.DataFrame) -> str:
    if pd.isna(emp_id) or emp_id is None or designation_df.empty:
        return ""
    if "employee_id" not in designation_df.columns:
        return ""
    row = designation_df[designation_df["employee_id"] == str(emp_id)]
    if not row.empty:
        return row.iloc[0]["designation"]
    return ""


def map_employee_branch(emp_id: Any, employee_branch_df: pd.DataFrame) -> str:
    if pd.isna(emp_id) or emp_id is None or employee_branch_df.empty:
        return ""
    if "employee_id" not in employee_branch_df.columns:
        return ""
    row = employee_branch_df[employee_branch_df["employee_id"] == str(emp_id)]
    if not row.empty:
        return row.iloc[0]["branch_name"]
    return ""


def map_employee_name(emp_id: Any, employee_name_df: pd.DataFrame) -> str:
    if pd.isna(emp_id) or emp_id is None or employee_name_df.empty:
        return ""
    if "employee_id" not in employee_name_df.columns:
        return ""
    row = employee_name_df[employee_name_df["employee_id"] == str(emp_id)]
    if not row.empty:
        return row.iloc[0]["employee_name"]
    return ""


# ==============================================================================
# Time and Shift Logic
# ==============================================================================


def _to_time(obj: Any) -> time | None:
    if pd.isna(obj) or obj is None or str(obj) == "":
        return None
    try:
        if hasattr(obj, "time") and hasattr(obj, "date"):
            return obj.time()
        if hasattr(obj, "hour") and hasattr(obj, "minute") and not hasattr(obj, "date"):
            return time(int(obj.hour), int(obj.minute))
        s = str(obj)
        s = s.strip()
        if " " in s:
            s = s.split(" ")[1]
        parts = s.split(":")
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        return time(hour, minute)
    except Exception:
        return None


def determine_shift_flag(start_time: Any, end_time: Any, shift_start: Any, shift_end: Any) -> str:
    flag = "normal"
    try:
        s_time = _to_time(start_time)
        e_time = _to_time(end_time)
        sh_start = _to_time(shift_start)
        sh_end = _to_time(shift_end)

        today = datetime.today().date()

        sh_start_dt = datetime.combine(today, sh_start) if sh_start else None
        sh_end_dt = datetime.combine(today, sh_end) if sh_end else None
        if sh_start_dt and sh_end_dt and sh_end_dt <= sh_start_dt:
            sh_end_dt = sh_end_dt + timedelta(days=1)

        def _to_dt_with_shift(t: time | None) -> datetime | None:
            if not t or not sh_start_dt:
                return datetime.combine(today, t) if t else None
            dt = datetime.combine(today, t)
            if sh_end_dt and sh_end_dt.date() != sh_start_dt.date() and dt < sh_start_dt:
                dt = dt + timedelta(days=1)
            return dt

        if s_time and sh_start:
            try:
                s_dt = _to_dt_with_shift(s_time)
                late_threshold = sh_start_dt + timedelta(minutes=5)
                if s_dt and s_dt > late_threshold:
                    flag = "late in"
            except Exception:
                try:
                    late_in_threshold = (datetime.combine(today, sh_start) + timedelta(minutes=5)).time()
                    if s_time > late_in_threshold:
                        flag = "late in"
                except Exception:
                    pass

        if e_time and sh_end:
            try:
                e_dt = _to_dt_with_shift(e_time)
                if (sh_start_dt and e_dt < sh_start_dt) or e_dt < sh_end_dt:
                    flag = "early out"
                else:
                    try:
                        grace_minutes = int(get_setting("late_checkout_grace_minutes") or 15)
                    except Exception:
                        grace_minutes = 15
                    normal_threshold_dt = sh_end_dt + timedelta(minutes=grace_minutes)
                    if e_dt <= normal_threshold_dt:
                        if flag != "late in":
                            flag = "normal"
                    else:
                        flag = "late checkout"
            except Exception:
                try:
                    if sh_start and e_time < sh_start:
                        flag = "early out"
                    elif e_time < sh_end:
                        flag = "early out"
                    else:
                        try:
                            grace_minutes = int(get_setting("late_checkout_grace_minutes") or 15)
                        except Exception:
                            grace_minutes = 15
                        normal_out_threshold = (datetime.combine(today, sh_end) + timedelta(minutes=grace_minutes)).time()
                        try:
                            if e_time <= normal_out_threshold:
                                if flag != "late in":
                                    flag = "normal"
                            else:
                                flag = "late checkout"
                        except Exception:
                            flag = "overtime"
                except Exception:
                    if e_time < sh_end:
                        flag = "early out"
                    else:
                        try:
                            grace_minutes = int(get_setting("late_checkout_grace_minutes") or 15)
                        except Exception:
                            grace_minutes = 15
                        normal_out_threshold = (datetime.combine(today, sh_end) + timedelta(minutes=grace_minutes)).time()
                        try:
                            if e_time <= normal_out_threshold:
                                if flag != "late in":
                                    flag = "normal"
                            else:
                                flag = "overtime"
                        except Exception:
                            flag = "overtime"
                else:
                    try:
                        grace_minutes = int(get_setting("late_checkout_grace_minutes") or 15)
                    except Exception:
                        grace_minutes = 15
                    normal_out_threshold = (datetime.combine(today, sh_end) + timedelta(minutes=grace_minutes)).time()
                    try:
                        if e_time <= normal_out_threshold:
                            if flag != "late in":
                                flag = "normal"
                        else:
                            flag = "late checkout"
                    except Exception:
                        flag = "overtime"
    except Exception:
        pass
    return flag


def determine_no_shift_flag(end_time: Any) -> str:
    flag = "normal"
    try:
        if pd.notna(end_time) and str(end_time) != "":
            end_time_obj = _to_time(end_time)
            if end_time_obj:
                if end_time_obj >= time(12, 0):
                    flag = "overtime"
                else:
                    flag = "early out"
    except Exception:
        pass
    return flag


def get_shift_info_with_capped(emp_id: str, work_status: str, start_time: Any, end_time: Any, shift_df: pd.DataFrame) -> tuple[str, str]:
    chosen_shift_name = ""
    chosen_shift_start = None
    chosen_shift_end = None

    if not shift_df.empty:
        shift_row = shift_df[shift_df["user_id"] == str(emp_id)]
        if not shift_row.empty:
            chosen_shift_name = shift_row.iloc[0]["shift_name"]
            chosen_shift_start = shift_row.iloc[0]["shift_start"]
            chosen_shift_end = shift_row.iloc[0]["shift_end"]

    if not chosen_shift_start or not chosen_shift_end:
        default_shift_name = get_default_shift() or ""
        if default_shift_name:
            try:
                templates = get_shift_templates() or []
                for t in templates:
                    if t.get("shift_name") == default_shift_name:
                        chosen_shift_name = t.get("shift_name")
                        chosen_shift_start = t.get("shift_start")
                        chosen_shift_end = t.get("shift_end")
                        break
            except Exception:
                pass

    if not chosen_shift_start or not chosen_shift_end:
        chosen_shift_name = chosen_shift_name or "Default"
        chosen_shift_start = chosen_shift_start or time(8, 0)
        chosen_shift_end = chosen_shift_end or time(17, 30)

    if work_status == "absent":
        return chosen_shift_name, "absent"

    if (pd.isna(end_time) or end_time is None or str(end_time) == "") and pd.notna(start_time):
        try:
            if hasattr(start_time, "date"):
                date_part = start_time.date()
            else:
                date_part = datetime.today().date()

            sh_end = _to_time(chosen_shift_end)
            if sh_end:
                shift_end_dt = datetime.combine(date_part, sh_end)
                try:
                    grace_minutes = int(get_setting("late_checkout_grace_minutes") or 15)
                except Exception:
                    grace_minutes = 15

                try:
                    cap_hours = int(get_setting("shift_cap_hours") or 8)
                except Exception:
                    cap_hours = 8

                grace_adjusted_end = shift_end_dt + timedelta(minutes=grace_minutes)
                cap_dt = grace_adjusted_end + timedelta(hours=cap_hours)
                if datetime.now() >= cap_dt:
                    return chosen_shift_name, "no checkout"
        except Exception:
            pass

    flag = determine_shift_flag(start_time, end_time, chosen_shift_start, chosen_shift_end)
    return chosen_shift_name, flag


# ==============================================================================
# Data Processing and Mapping Application
# ==============================================================================


def apply_branch_mappings(summary_df: pd.DataFrame) -> pd.DataFrame:
    mappings = get_device_branch_mappings() or []
    mappings_df = pd.DataFrame(mappings)
    if mappings_df is None or mappings_df.empty:
        summary_df["start_device_sn_branch"] = ""
        summary_df["end_device_sn_branch"] = ""
        return summary_df

    summary_df["start_device_sn_branch"] = summary_df.apply(
        lambda row: map_branch(row.get("start_device_sn"), mappings_df) if row.get("work_status") == "worked" else "",
        axis=1,
    )
    summary_df["end_device_sn_branch"] = summary_df.apply(
        lambda row: map_branch(row.get("end_device_sn"), mappings_df) if row.get("work_status") == "worked" else "",
        axis=1,
    )
    return summary_df


def apply_designation_mappings(summary_df: pd.DataFrame) -> pd.DataFrame:
    designation_mappings = get_employee_designation_mappings() or []
    designation_df = pd.DataFrame(designation_mappings)
    if not summary_df.empty and "employee_id" in summary_df.columns:
        summary_df["designation"] = summary_df["employee_id"].apply(lambda emp_id: map_designation(emp_id, designation_df))
    else:
        summary_df["designation"] = ""
    return summary_df


def apply_employee_branch_mappings(summary_df: pd.DataFrame) -> pd.DataFrame:
    employee_branch_mappings = get_employee_branch_mappings() or []
    employee_branch_df = pd.DataFrame(employee_branch_mappings)
    if not summary_df.empty and "employee_id" in summary_df.columns:
        summary_df["employee_branch"] = summary_df["employee_id"].apply(lambda emp_id: map_employee_branch(emp_id, employee_branch_df))
    else:
        summary_df["employee_branch"] = ""
    return summary_df


def apply_employee_name_mappings(summary_df: pd.DataFrame) -> pd.DataFrame:
    employee_name_mappings = get_employee_name_mappings() or []
    employee_name_df = pd.DataFrame(employee_name_mappings)
    if not summary_df.empty and "employee_id" in summary_df.columns:
        summary_df["employee_name"] = summary_df["employee_id"].apply(lambda emp_id: map_employee_name(emp_id, employee_name_df))
    else:
        summary_df["employee_name"] = ""
    return summary_df


def apply_shift_mappings(summary_df: pd.DataFrame, shift_mappings: list[dict[str, Any]] | None) -> pd.DataFrame:
    shift_df = pd.DataFrame(shift_mappings) if shift_mappings else pd.DataFrame()

    summary_df["shift_name"] = ""
    summary_df["shift_flag"] = ""
    summary_df["late_in"] = False
    summary_df["late_in_time"] = ""

    for idx, row in summary_df.iterrows():
        no_checkout = row.get("no_checkout", False)

        if no_checkout:
            shift_name, _ = get_shift_info_with_capped(row["employee_id"], row["work_status"], row["start_time"], row["end_time"], shift_df)
            flag = "no checkout"
        else:
            shift_name, flag = get_shift_info_with_capped(row["employee_id"], row["work_status"], row["start_time"], row["end_time"], shift_df)

        try:
            if not no_checkout and bool(row.get("early_checkout", False)):
                flag = "early out"
        except Exception:
            pass

        summary_df.loc[idx, "shift_name"] = shift_name
        summary_df.loc[idx, "shift_flag"] = flag

        try:
            s_time = _to_time(row.get("start_time"))
            sh_start = _to_time(
                shift_name and shift_df[shift_df["user_id"] == str(row["employee_id"])].iloc[0]["shift_start"]
                if not shift_df.empty and not shift_df[shift_df["user_id"] == str(row["employee_id"])].empty
                else shift_name and None
            )
        except Exception:
            s_time = None
            sh_start = None

        if s_time is None or sh_start is None:
            try:
                if shift_name:
                    templates = get_shift_templates() or []
                    for t in templates:
                        if t.get("shift_name") == shift_name:
                            sh_start = _to_time(t.get("shift_start"))
                            break
            except Exception:
                pass

        late_in_flag = False
        if s_time and sh_start:
            try:
                today = datetime.today().date()
                s_dt = datetime.combine(today, s_time)
                late_threshold_dt = datetime.combine(today, sh_start) + timedelta(minutes=5)
                if s_dt > late_threshold_dt:
                    late_in_flag = True
            except Exception:
                late_in_flag = False

        summary_df.loc[idx, "late_in"] = late_in_flag
        if late_in_flag:
            try:
                summary_df.loc[idx, "late_in_time"] = s_time.strftime("%H:%M:%S") if s_time else str(row.get("start_time") or "")
            except Exception:
                summary_df.loc[idx, "late_in_time"] = str(row.get("start_time") or "")

    return summary_df


def clean_attendance_summary(summary_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df.empty:
        return summary_df

    if "start_time" in summary_df.columns:
        summary_df["start_time"] = summary_df["start_time"].apply(lambda x: x.strftime("%H:%M:%S") if pd.notna(x) and hasattr(x, "strftime") else "")

    if "end_time" in summary_df.columns:
        summary_df["end_time"] = summary_df["end_time"].apply(lambda x: x.strftime("%H:%M:%S") if pd.notna(x) and hasattr(x, "strftime") else "")

    columns_to_remove = ["start_device_sn", "end_device_sn", "no_checkout", "designation", "employee_branch"]

    for col in columns_to_remove:
        if col in summary_df.columns:
            summary_df = summary_df.drop(columns=[col])

    return summary_df


def create_subtotal_rows(summary_df: pd.DataFrame) -> list[dict[str, Any]]:
    output_rows = []

    for emp_id, group in summary_df.groupby("employee_id", sort=False):
        output_rows.extend(group.to_dict(orient="records"))

        worked_group = group[group["work_status"] == "worked"].copy()
        days_worked = len(worked_group)

        if not worked_group.empty:
            worked_group.loc[:, "time_spent_td"] = pd.to_timedelta(worked_group["time_spent"])
            subtotal = worked_group["time_spent_td"].sum()
            subtotal_str = str(subtotal).split(".")[0]
        else:
            subtotal_str = "0:00:00"

        subtotal_row = dict.fromkeys(summary_df.columns, "")
        subtotal_row["employee_id"] = emp_id
        if not group.empty:
            subtotal_row["employee_name"] = group.iloc[0].get("employee_name", "")
        subtotal_row["day"] = "Subtotal"
        subtotal_row["days_worked"] = days_worked
        subtotal_row["total_hours"] = subtotal_str
        subtotal_row["work_status"] = "subtotal"
        output_rows.append(subtotal_row)

    return output_rows


# ==============================================================================
# Main Report Generation
# ==============================================================================


def generate_attendance_summary(
    attendences: list[dict[str, Any]],
    _device_logs: list[dict[str, Any]],
    _finger_logs: list[dict[str, Any]],
    _migration_logs: list[dict[str, Any]],
    _user_logs: list[dict[str, Any]],
    shift_mappings: list[dict[str, Any]] | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    branch_name: str | None = None,
    employee_branch: str | None = None,
    employee_name: str | None = None,
    designation: str | None = None,
) -> pd.DataFrame:
    summary_df = process_attendance_summary(attendences, start_date, end_date, branch_name, employee_branch, employee_name, designation)

    if summary_df is not None and not summary_df.empty:
        summary_df = apply_branch_mappings(summary_df)
        summary_df = apply_designation_mappings(summary_df)
        summary_df = apply_employee_branch_mappings(summary_df)
        summary_df = apply_employee_name_mappings(summary_df)
        summary_df = apply_shift_mappings(summary_df, shift_mappings)
        summary_df = clean_attendance_summary(summary_df)
        output_rows = create_subtotal_rows(summary_df)
        merged = pd.DataFrame(output_rows, columns=[*summary_df.columns.tolist(), "days_worked", "total_hours"])
    else:
        merged = summary_df if summary_df is not None else pd.DataFrame()

    return merged


def generate_branch_attendance_summary(
    attendences: list[dict[str, Any]],
    branch_employee_ids: list[str],
    _device_logs: list[dict[str, Any]],
    _finger_logs: list[dict[str, Any]],
    _migration_logs: list[dict[str, Any]],
    _user_logs: list[dict[str, Any]],
    shift_mappings: list[dict[str, Any]] | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
) -> pd.DataFrame:
    if not start_date or not end_date or not branch_employee_ids:
        return pd.DataFrame()

    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    all_dates = []
    current_date = start_dt
    while current_date <= end_dt:
        all_dates.append(current_date.strftime("%Y-%m-%d"))
        current_date += timedelta(days=1)

    all_records = []
    existing_summary = pd.DataFrame()
    if attendences:
        existing_summary = process_attendance_summary(attendences, start_date, end_date, None, None, None, None)
        if existing_summary is not None and not existing_summary.empty:
            existing_summary = apply_branch_mappings(existing_summary)
            existing_summary = apply_designation_mappings(existing_summary)
            existing_summary = apply_employee_branch_mappings(existing_summary)
            existing_summary = apply_employee_name_mappings(existing_summary)
            existing_summary = apply_shift_mappings(existing_summary, shift_mappings)
            existing_summary = clean_attendance_summary(existing_summary)

    existing_lookup = {}
    if not existing_summary.empty:
        for _, row in existing_summary.iterrows():
            key = f"{row['employee_id']}_{row['day']}"
            existing_lookup[key] = row.to_dict()

    for emp_id in branch_employee_ids:
        for date in all_dates:
            key = f"{emp_id}_{date}"
            if key in existing_lookup:
                all_records.append(existing_lookup[key])
            else:
                absent_record = {
                    "employee_id": emp_id,
                    "day": date,
                    "start_time": "",
                    "end_time": "",
                    "start_device_sn_branch": "",
                    "end_device_sn_branch": "",
                    "time_spent": "0:00:00",
                    "work_status": "absent",
                    "shift_name": "",
                    "shift_flag": "absent",
                    "late_in": False,
                    "late_in_time": "",
                    "employee_name": "",
                    "employee_branch": "",
                }
                all_records.append(absent_record)

    if all_records:
        complete_summary = pd.DataFrame(all_records)
        complete_summary = apply_designation_mappings(complete_summary)
        complete_summary = apply_employee_branch_mappings(complete_summary)
        complete_summary = apply_employee_name_mappings(complete_summary)
        output_rows = create_subtotal_rows(complete_summary)
        merged = pd.DataFrame(output_rows, columns=[*complete_summary.columns.tolist(), "days_worked", "total_hours"])
    else:
        merged = pd.DataFrame()

    return merged


# ==============================================================================
# Excel Highlighting and Styling
# ==============================================================================


def find_column_indices(ws: Any) -> tuple[int | None, int | None, int | None]:
    day_col = None
    work_status_col = None
    no_checkout_col = None

    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        header = str(cell.value).strip().lower()
        if header in ("day", "date"):
            day_col = idx
        if header in ("work_status", "work status", "workstatus"):
            work_status_col = idx
        if header in ("no_checkout", "no checkout", "shift capped", "shiftcap"):
            no_checkout_col = idx

    return day_col, work_status_col, no_checkout_col


def apply_subtotal_highlighting(row: list, day_col: int | None, blue_fill: PatternFill) -> None:
    if not day_col:
        return
    try:
        cell_val = row[day_col - 1].value
        if cell_val is None:
            return
        if str(cell_val).strip().lower() == "subtotal":
            for cell in row:
                cell.fill = blue_fill
    except Exception:
        return


def apply_status_highlighting(row: list, work_status_col: int | None, day_col: int | None, green_fill: PatternFill, orange_fill: PatternFill, pink_fill: PatternFill) -> None:
    if not work_status_col:
        return

    is_sunday = False
    if day_col:
        try:
            day_val = row[day_col - 1].value
            if day_val:
                if isinstance(day_val, str):
                    try:
                        date_obj = datetime.strptime(day_val, "%Y-%m-%d")
                        if date_obj.weekday() == 6:
                            is_sunday = True
                    except Exception:
                        pass
                elif hasattr(day_val, 'weekday') and day_val.weekday() == 6:
                    is_sunday = True
        except Exception:
            pass

    if is_sunday:
        for cell in row:
            cell.fill = pink_fill
        return

    work_status = row[work_status_col - 1].value

    if work_status == "worked":
        for cell in row:
            cell.fill = green_fill
    elif work_status == "absent":
        for cell in row:
            cell.fill = orange_fill


def apply_row_highlighting(ws: Any) -> None:
    day_col, work_status_col, no_checkout_col = find_column_indices(ws)

    blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FFD7A3", end_color="FFD7A3", fill_type="solid")
    pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Pink for Sundays

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        apply_subtotal_highlighting(row, day_col, blue_fill)
        apply_status_highlighting(row, work_status_col, day_col, green_fill, light_orange_fill, pink_fill)


def apply_flag_highlighting(ws: Any) -> None:
    flag_col_idx = None
    work_status_col_idx = None
    day_col_idx = None

    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            header = str(cell.value).strip().lower()
            if header in ("shift flag", "shift_flag", "shiftflag"):
                flag_col_idx = idx
            elif header in ("work_status", "work status", "workstatus"):
                work_status_col_idx = idx
            elif header in ("day", "date"):
                day_col_idx = idx

    if not flag_col_idx:
        return

    overtime_fill = PatternFill(start_color="E6E0FF", end_color="E6E0FF", fill_type="solid")  # Light purple for overtime
    worked_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for all other worked statuses
    pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Pink for Sundays

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        try:
            # Check if this is a Sunday first (highest priority after subtotal)
            is_sunday = False
            if day_col_idx:
                try:
                    day_val = row[day_col_idx - 1].value
                    if day_val:
                        if isinstance(day_val, str):
                            if day_val.strip().lower() == "subtotal":
                                continue
                            try:
                                date_obj = datetime.strptime(day_val, "%Y-%m-%d")
                                if date_obj.weekday() == 6:  # Sunday
                                    is_sunday = True
                            except Exception:
                                pass
                        elif hasattr(day_val, 'weekday') and day_val.weekday() == 6:
                            is_sunday = True
                except Exception:
                    pass

            if work_status_col_idx:
                work_status = str(row[work_status_col_idx - 1].value or "").strip().lower()
                if work_status in ("absent", "subtotal"):
                    continue

            if is_sunday:
                for c in row:
                    c.fill = pink_fill
                continue

            cell = row[flag_col_idx - 1]
            val_raw = cell.value or ""
            val = str(val_raw).strip().lower()

            has_overtime = "overtime" in val or "over time" in val

            if has_overtime:
                fill = overtime_fill
            else:
                # All other worked statuses (normal, late in, late checkout, early out, no checkout) get green
                fill = worked_fill

            for c in row:
                c.fill = fill
        except Exception:
            continue


# ==============================================================================
# Employee Summary Sheet Generation
# ==============================================================================


def create_employee_summary_sheet(summary_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df.empty:
        return pd.DataFrame()

    summary_rows = []

    for emp_id, group in summary_df.groupby("employee_id", sort=False):
        first_row = group.iloc[0]

        total_days = len(group)
        worked_days = len(group[group["work_status"] == "worked"])
        absent_days = len(group[group["work_status"] == "absent"])

        worked_group = group[group["work_status"] == "worked"].copy()
        if not worked_group.empty:
            worked_group.loc[:, "time_spent_td"] = pd.to_timedelta(worked_group["time_spent"])
            subtotal = worked_group["time_spent_td"].sum()
            subtotal_str = str(subtotal).split(".")[0]
            avg_hours_per_day = subtotal / worked_days if worked_days > 0 else pd.Timedelta(0)
            avg_hours_str = str(avg_hours_per_day).split(".")[0]
        else:
            subtotal_str = "0:00:00"
            avg_hours_str = "0:00:00"

        shift_name = first_row.get("shift_name", "")
        worked_mask = group["work_status"].astype(str).str.lower() == "worked"

        flag_series = group.get("shift_flag", pd.Series([""] * len(group), index=group.index))
        flag_norm = (
            flag_series.astype(str)
            .str.strip()
            .str.lower()
            .replace(
                {
                    "shift capped": "no checkout",
                    "shift cap": "no checkout",
                    "shiftcap": "no checkout",
                    "latein": "late in",
                    "earlyin": "early in",
                    "earlyout": "early out",
                    "over time": "overtime",
                }
            )
        )

        late_in_count = int(group.loc[worked_mask, "late_in"].fillna(False).astype(bool).sum()) if "late_in" in group.columns else int((flag_norm.eq("late in") & worked_mask).sum())
        early_out_count = int((flag_norm.eq("early out") & worked_mask).sum())
        late_checkout_count = int((flag_norm.isin(["late checkout", "overtime"]) & worked_mask).sum())
        on_time_count = int((flag_norm.isin(["on time", "normal"]) & worked_mask).sum())

        if "no_checkout" in group.columns:
            try:
                no_checkout_count = int(group.loc[worked_mask, "no_checkout"].fillna(value=False).astype(bool).sum())
            except Exception:
                no_checkout_count = int(sum(1 for v in group.loc[worked_mask, "no_checkout"] if bool(v)))
        else:
            no_checkout_count = int((flag_norm.eq("no checkout") & worked_mask).sum())

        attendance_percentage = (worked_days / total_days * 100) if total_days > 0 else 0

        summary_row = {
            "employee_id": emp_id,
            "employee_name": first_row.get("employee_name", ""),
            "shift_name": shift_name,
            "total_days": total_days,
            "days_worked": worked_days,
            "days_absent": absent_days,
            "attendance_percentage": f"{attendance_percentage:.1f}%",
            "total_hours_worked": subtotal_str,
            "avg_hours_per_day": avg_hours_str,
            "on_time_days": on_time_count,
            "late_in_days": late_in_count,
            "early_out_days": early_out_count,
            "late_checkout_days": late_checkout_count,
            "no_checkout_days": no_checkout_count,
        }
        summary_rows.append(summary_row)

    summary_summary_df = pd.DataFrame(summary_rows)
    if not summary_summary_df.empty:
        summary_summary_df["employee_id"] = summary_summary_df["employee_id"].astype(str)
        summary_summary_df = summary_summary_df.sort_values("employee_id")

    return summary_summary_df


# ==============================================================================
# File Export Functions
# ==============================================================================


def write_csv(
    attendences: list[dict[str, Any]],
    device_logs: list[dict[str, Any]],
    finger_logs: list[dict[str, Any]],
    migration_logs: list[dict[str, Any]],
    user_logs: list[dict[str, Any]],
    merged: pd.DataFrame,
) -> io.BytesIO:
    output = io.BytesIO()
    employee_summary = create_employee_summary_sheet(merged)

    export_df_columns = {
        "EPF Number": "employee_id",
        "Employee Name": "employee_name",
        "Date": "day",
        "In Time": "start_time",
        "Out Time": "end_time",
        "Working Hours": "time_spent",
        "Work Status": "work_status",
        "In Location": "start_device_sn_branch",
        "Out Location": "end_device_sn_branch",
        "Shift Flag": "shift_flag",
        "Total Work Dates": "days_worked",
        "Total Work Hours": "total_hours",
    }

    if merged is None:
        export_df = pd.DataFrame(columns=export_df_columns.keys())
    else:
        export_df = pd.DataFrame()
        for new_col, old_col in export_df_columns.items():
            export_df[new_col] = merged[old_col] if old_col in merged.columns else ""

        try:
            late_series = merged["late_in"].fillna(False).astype(bool) if "late_in" in merged.columns else pd.Series([False] * len(merged))
            base_flags = export_df["Shift Flag"].fillna("").astype(str)
            combined_flags = []
            for idx, bf in base_flags.items():
                parts = [p.strip() for p in str(bf).split(";") if p.strip()]
                if late_series.iloc[idx] and "late in" not in [p.lower() for p in parts]:
                    parts.append("late in")
                combined_flags.append("; ".join(parts))
            export_df["Shift Flag"] = combined_flags
        except Exception:
            pass

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        employee_summary.to_excel(writer, sheet_name="Employee Summary", index=False)
        export_df.to_excel(writer, sheet_name="Attendance Details", index=False)
        ws_details = writer.sheets["Attendance Details"]
        apply_row_highlighting(ws_details)
        apply_flag_highlighting(ws_details)

    output.seek(0)
    return output


def write_branch_summary_csv(merged: pd.DataFrame, branch_name: str) -> io.BytesIO:
    output = io.BytesIO()

    export_df_columns = {
        "EPF Number": "employee_id",
        "Employee Name": "employee_name",
        "Date": "day",
        "In Time": "start_time",
        "Out Time": "end_time",
        "Working Hours": "time_spent",
        "Work Status": "work_status",
        "In Location": "start_device_sn_branch",
        "Out Location": "end_device_sn_branch",
        "Shift Flag": "shift_flag",
        "Total Work Dates": "days_worked",
        "Total Work Hours": "total_hours",
    }

    if merged is None or merged.empty:
        export_df = pd.DataFrame(columns=export_df_columns.keys())
    else:
        export_df = pd.DataFrame()
        for new_col, old_col in export_df_columns.items():
            export_df[new_col] = merged[old_col] if old_col in merged.columns else ""

        try:
            late_series = merged["late_in"].fillna(False).astype(bool) if "late_in" in merged.columns else pd.Series([False] * len(merged))
            base_flags = export_df["Shift Flag"].fillna("").astype(str)
            combined_flags = []
            for idx, bf in base_flags.items():
                parts = [p.strip() for p in str(bf).split(";") if p.strip()]
                if late_series.iloc[idx] and "late in" not in [p.lower() for p in parts]:
                    parts.append("late in")
                combined_flags.append("; ".join(parts))
            export_df["Shift Flag"] = combined_flags
        except Exception:
            pass

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name=f"{branch_name} Summary", index=False)
        ws = writer.sheets[f"{branch_name} Summary"]
        apply_row_highlighting(ws)
        apply_flag_highlighting(ws)

    output.seek(0)
    return output
